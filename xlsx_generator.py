"""XLSX Generator — creates a pricing model spreadsheet from BOM data.

Exactly matches the Prologis pricing model format (Wesco reference).
Single "Pricing" sheet. No fills, no borders except thin-bottom on header row 1.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Number formats (exact match from Wesco)
ACCT_MONEY = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
ACCT_WHOLE = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
ACCT_NUM   = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
ACCT_DEC   = '_(* #,##0.0_);_(* \\(#,##0.0\\);_(* "-"??_);_(@_)'
ACCT_DEC2  = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
PCT_FMT    = '0.0%'
DOLLAR_DEC = '"$"#,##0.00'

THIN_BOTTOM = Border(bottom=Side(style="thin"))
FONT_11     = Font(name="Calibri", size=11)
FONT_11B    = Font(name="Calibri", size=11, bold=True)


def _c(ws, row, col, value=None, bold=False, fmt=None, halign=None):
    """Helper to set a cell with exact styling."""
    # Never write empty strings — they create invalid inlineStr XML in openpyxl
    if value == "":
        value = None
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_11B if bold else FONT_11
    if fmt:
        cell.number_format = fmt
    if halign:
        cell.alignment = Alignment(horizontal=halign)
    return cell


def generate_pricing_xlsx(project: dict, bay_types: list, bom: list) -> bytes:
    """Generate a pricing model XLSX matching the Prologis format exactly."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing"

    # Column widths (from Wesco)
    widths = {
        "A": 46.83, "B": 20.83, "C": 29.16, "D": 18.5, "E": 21.83,
        "F": 26.16, "G": 13.5, "H": 18.5, "I": 9.0,
        "J": 36.5, "K": 20.16, "L": 18.16, "M": 13.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # === Row 1: Headers — bold, thin bottom border, H centered ===
    for ci, h in enumerate(["Item", "QTY", "MFG", "Cost", "Price", "Total Cost", "Total Price", "%"], 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = FONT_11B
        cell.border = THIN_BOTTOM
        if ci == 8:  # % column
            cell.alignment = Alignment(horizontal="center")

    # === Row 2: "Materials:" ===
    _c(ws, 2, 1, "Materials:", bold=True)

    # Row 3: empty but has % formula in H
    _c(ws, 3, 8, '=IFERROR((E3-D3)/E3,"")', fmt=PCT_FMT)

    # === Sidebar: J3/K3 — Project Margin ===
    _c(ws, 3, 10, "Project Margin")
    _c(ws, 3, 11, 0, fmt=PCT_FMT)

    row = 4

    # === Material line items (start at row 4) ===
    first_mat_row = row
    for item in bom:
        _c(ws, row, 1, item.get("description", ""))
        _c(ws, row, 2, item.get("total_qty", 0))
        _c(ws, row, 3, item.get("mfg", ""))
        _c(ws, row, 4, 0, fmt=ACCT_MONEY)
        _c(ws, row, 5, f'=ROUND(D{row}/(1-$K$3),2)', fmt=ACCT_MONEY)
        _c(ws, row, 6, f'=D{row}*B{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 7, f'=B{row}*E{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
        row += 1
    last_mat_row = row - 1

    # === Sidebar: Pricing Summary (rows 6-13, positioned relative to header) ===
    _c(ws, 6, 10, "Pricing Summary", bold=True, halign="center")
    _c(ws, 7, 11, "Domestic", bold=True)

    # These row refs will be filled after we know install/freight/services rows
    # We'll come back and write them at the end

    # Blank row
    _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
    row += 1

    # === Install section ===
    install_label_row = row
    _c(ws, row, 1, "Install:", bold=True, halign="left")
    row += 1
    install_start = row

    for name in ["Main Scope", "Lift Rental"]:
        _c(ws, row, 1, name)
        _c(ws, row, 2, 1)
        _c(ws, row, 3, "")
        _c(ws, row, 4, 0, fmt=ACCT_MONEY)
        _c(ws, row, 5, f'=ROUND(D{row}/(1-$K$3),2)', fmt=ACCT_MONEY)
        _c(ws, row, 6, f'=D{row}*B{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 7, f'=B{row}*E{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
        row += 1
    install_end = row - 1

    # Blank rows with % formulas (matching Wesco pattern)
    _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
    row += 1
    _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
    row += 1

    # === Freight section ===
    freight_label_row = row
    _c(ws, row, 1, "Freight", bold=True, halign="left")
    _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
    row += 1
    freight_start = row

    mfg_name = project.get("manufacturer", "") or "Rack"
    freight_items = [mfg_name, "Hilti", "WWMH"]
    freight_mfgs  = [f"{mfg_name} Freight", "Hilti Freight", "WWMH Freight"]
    for name, mfg in zip(freight_items, freight_mfgs):
        _c(ws, row, 1, name, halign="left")
        _c(ws, row, 2, 1, fmt=ACCT_DEC2)
        _c(ws, row, 3, mfg)
        _c(ws, row, 4, 0, fmt=ACCT_MONEY)
        _c(ws, row, 5, f'=ROUND(D{row}/(1-$K$3),2)', fmt=ACCT_MONEY)
        _c(ws, row, 6, f'=D{row}*B{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 7, f'=B{row}*E{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
        row += 1
    freight_end = row - 1

    # Blank row with price formula
    _c(ws, row, 5, f'=ROUND(D{row}/(1-$K$3),2)', fmt=ACCT_MONEY)
    row += 1

    # === Services section (no section label — items listed directly) ===
    # In Wesco: PM, TCO, High Pile, Permit Services, Eng Calc, Dumpsters
    svc_items = [
        ("Project Management", 0, True),      # QTY editable, has cost & price cols
        ("TCO & Project Uncertainties", 0, True),
        ("High Pile", 1, True),
        ("Permit Services", 1, True),
        ("Engineering Calculations", 1, True),
        ("Dumpsters", 1, True),
    ]
    svc_start = row
    svc_rows = {}
    for name, default_qty, _ in svc_items:
        _c(ws, row, 1, name, halign="left")
        _c(ws, row, 2, default_qty, fmt=ACCT_DEC)
        _c(ws, row, 4, 0, fmt=ACCT_WHOLE)
        _c(ws, row, 5, 0, fmt=ACCT_WHOLE)
        _c(ws, row, 6, f'=D{row}*B{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 7, f'=B{row}*E{row}', fmt=ACCT_WHOLE)
        _c(ws, row, 8, f'=IFERROR((E{row}-D{row})/E{row},"")', fmt=PCT_FMT)
        svc_rows[name] = row
        row += 1
    svc_end = row - 1

    pm_row = svc_rows["Project Management"]
    tco_row = svc_rows["TCO & Project Uncertainties"]
    hp_row = svc_rows["High Pile"]
    permit_row = svc_rows["Permit Services"]
    eng_row = svc_rows["Engineering Calculations"]
    dump_row = svc_rows["Dumpsters"]

    # === Grand Total ===
    total_row = row
    _c(ws, row, 1, "Grand Total", bold=True, halign="left")
    _c(ws, row, 6, f'=SUM(F3:F{svc_end})', bold=True, fmt=ACCT_WHOLE)
    _c(ws, row, 7, f'=SUM(G3:G{svc_end})', bold=True, fmt=ACCT_WHOLE)
    row += 1

    # === Profit | Margin ===
    profit_row = row
    _c(ws, row, 1, "Profit | Margin", bold=True, halign="left")
    _c(ws, row, 6, f'=G{total_row}-F{total_row}', fmt=ACCT_WHOLE)
    _c(ws, row, 7, f'=F{profit_row}/G{total_row}', fmt=PCT_FMT)
    row += 1

    # ==================== SIDEBAR ====================

    # --- Pricing Summary (J8-K13) ---
    _c(ws, 8, 10, "Rack Material")
    _c(ws, 8, 11, f'=SUM(G{first_mat_row}:G{last_mat_row})', fmt=ACCT_WHOLE)

    _c(ws, 9, 10, "Installation")
    _c(ws, 9, 11, f'=SUM(G{install_start}:G{install_end})', fmt=ACCT_WHOLE)

    _c(ws, 10, 10, "Freight")
    _c(ws, 10, 11, f'=SUM(G{freight_start}:G{freight_end})', fmt=ACCT_WHOLE, halign="right")

    _c(ws, 11, 10, "Project Management & Permit Services")
    _c(ws, 11, 11, f'=G{pm_row}+G{permit_row}+G{tco_row}+G{dump_row}', fmt=ACCT_WHOLE)

    _c(ws, 12, 10, "Engineering Calculations & High Pile", halign="left")
    _c(ws, 12, 11, f'=G{eng_row}+G{hp_row}', fmt=ACCT_WHOLE)

    _c(ws, 13, 10, "Project Total", bold=True)
    _c(ws, 13, 11, '=SUM(K8:K12)', bold=True, fmt=ACCT_WHOLE)

    # --- Pallet Positions (row 21) ---
    pp = project.get("total_pallet_positions", 0)
    _c(ws, 21, 10, "Pallet Positions", bold=True)
    _c(ws, 21, 11, pp, fmt=ACCT_NUM)
    _c(ws, 21, 12, '=K13/K21', fmt=DOLLAR_DEC)

    # --- Comparison: Materials (starts 2 rows after pallet positions) ---
    cr = 26  # matching Wesco row position approximately — adjusts based on total_row
    # Place it relative to the main content
    comp_start = total_row + 4 if total_row + 4 > 26 else 26

    # Materials comparison
    _c(ws, comp_start, 10, "Materials", bold=True, halign="left")
    _c(ws, comp_start, 11, "Model", bold=True, fmt=ACCT_MONEY, halign="center")
    _c(ws, comp_start, 12, "Quote", bold=True, halign="center")

    mfgs_seen = []
    for item in bom:
        m = item.get("mfg", "")
        if m and m not in mfgs_seen:
            mfgs_seen.append(m)

    cr = comp_start + 1
    for m in mfgs_seen:
        _c(ws, cr, 10, m)
        _c(ws, cr, 11, f'=SUMIF($C:$C, J{cr}, $F:$F)', fmt=ACCT_WHOLE)
        _c(ws, cr, 12, 0, fmt=ACCT_WHOLE)
        _c(ws, cr, 13, f'=K{cr}-L{cr}', fmt=ACCT_WHOLE)
        cr += 1

    cr += 1
    _c(ws, cr, 10, "Total", bold=True)
    mat_comp_first = comp_start + 1
    mat_comp_last = cr - 2
    _c(ws, cr, 11, f'=SUM(K{mat_comp_first}:K{mat_comp_last})', bold=True, fmt=ACCT_WHOLE)
    _c(ws, cr, 12, f'=SUM(L{mat_comp_first}:L{mat_comp_last})', bold=True, fmt=ACCT_WHOLE)
    _c(ws, cr, 13, f'=L{cr}-K{cr}', bold=True, fmt=ACCT_MONEY)
    cr += 2

    # Freight comparison
    _c(ws, cr, 10, "Freight", bold=True, halign="left")
    _c(ws, cr, 11, "Model", bold=True, fmt=ACCT_MONEY, halign="center")
    _c(ws, cr, 12, "Quote", bold=True, halign="center")
    cr += 1
    frt_comp_first = cr
    for mfg_label in freight_mfgs:
        _c(ws, cr, 10, mfg_label)
        _c(ws, cr, 11, f'=SUMIF($C:$C, J{cr}, $F:$F)', fmt=ACCT_WHOLE)
        _c(ws, cr, 12, 0, fmt=ACCT_WHOLE)
        _c(ws, cr, 13, f'=L{cr}-K{cr}', fmt=ACCT_MONEY)
        cr += 1
    frt_comp_last = cr - 1

    cr += 1
    _c(ws, cr, 10, "Total", bold=True)
    _c(ws, cr, 11, f'=SUM(K{frt_comp_first}:K{frt_comp_last})', bold=True, fmt=ACCT_WHOLE)
    _c(ws, cr, 12, f'=SUM(L{frt_comp_first}:L{frt_comp_last})', bold=True, fmt=ACCT_WHOLE)
    _c(ws, cr, 13, f'=L{cr}-K{cr}', bold=True, fmt=ACCT_MONEY)
    cr += 2

    # Labor comparison
    _c(ws, cr, 10, "Labor", bold=True, halign="left")
    _c(ws, cr, 11, "Model", bold=True, fmt=ACCT_MONEY, halign="center")
    _c(ws, cr, 12, "Quote", bold=True, halign="center")

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
